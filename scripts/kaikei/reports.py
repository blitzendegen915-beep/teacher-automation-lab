"""xlsx帳票の生成。

- write_settlement_xlsx : 立替金精算表
- write_camp_report_xlsx : 保護者向け 収支報告（昨年様式を踏襲、3シート）
- write_headcount_xlsx  : 日×食事の人数表（ホテル連絡用）

すべて output/ に出力する。数値は極力Excel式（=SUM(...)等）で持たせ、
未確定の値は黄色セルとして残す。
"""
from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from . import model
from .checks import date_ja

ROOT = Path(__file__).resolve().parents[2]
OUTPUT_DIR = ROOT / "output"

YEN_FMT = "#,##0"
YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
HEADER_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
BOLD = Font(bold=True)
TITLE_FONT = Font(bold=True, size=14)
THIN = Side(style="thin", color="BBBBBB")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _pending_cell(ws, cell_ref, label="未確定"):
    cell = ws[cell_ref]
    cell.fill = YELLOW
    if cell.value in (None, ""):
        cell.value = 0
    cell.number_format = YEN_FMT
    return cell


# ---------------------------------------------------------------------------
# 立替金精算表
# ---------------------------------------------------------------------------


def write_settlement_xlsx(path: str | Path = None, camp: model.Camp = None) -> Path:
    if camp is None:
        camp = model.load()
    if path is None:
        path = OUTPUT_DIR / f"{camp.id}_立替金精算表.xlsx"
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    settle = model.settlement(camp)

    wb = Workbook()
    ws = wb.active
    ws.title = "立替金精算表"
    _autosize(ws, [12, 12, 24, 30, 12, 10, 30])

    ws["A1"] = f"{camp.name} 立替金精算表"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "作成日:"
    ws["B2"] = date.today()
    ws["B2"].number_format = "yyyy/mm/dd"

    headers = ["立替者", "日付", "支払先", "内容", "金額", "領収書", "備考"]
    header_row = 4
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=c, value=h)
        cell.font = BOLD
        cell.fill = HEADER_FILL
        cell.border = BOX

    row = header_row + 1
    payer_total_rows = []
    for payer, info in settle["by_payer"].items():
        first_row = row
        for e in info["items"]:
            ws.cell(row=row, column=1, value=payer)
            ws.cell(row=row, column=2, value=e.date).number_format = "m/d"
            ws.cell(row=row, column=3, value=e.vendor)
            ws.cell(row=row, column=4, value=e.description)
            amt = ws.cell(row=row, column=5, value=e.amount)
            amt.number_format = YEN_FMT
            ws.cell(row=row, column=6, value=e.receipt)
            ws.cell(row=row, column=7, value=e.note)
            for c in range(1, 8):
                ws.cell(row=row, column=c).border = BOX
            row += 1
        subtotal_row = row
        ws.cell(row=subtotal_row, column=4, value=f"{payer} 小計").font = BOLD
        subtotal_cell = ws.cell(
            row=subtotal_row, column=5, value=f"=SUM(E{first_row}:E{row - 1})"
        )
        subtotal_cell.font = BOLD
        subtotal_cell.number_format = YEN_FMT
        for c in range(1, 8):
            ws.cell(row=subtotal_row, column=c).border = BOX
        payer_total_rows.append(subtotal_row)
        row += 2

    grand_row = row
    ws.cell(row=grand_row, column=4, value="立替金 合計（各先生への返金額）").font = BOLD
    formula = "=" + "+".join(f"E{r}" for r in payer_total_rows)
    grand_cell = ws.cell(row=grand_row, column=5, value=formula)
    grand_cell.font = BOLD
    grand_cell.number_format = YEN_FMT
    grand_cell.fill = HEADER_FILL

    ws.cell(row=grand_row + 2, column=1, value="※ settled=no（未精算）の支出のみを対象としています。")
    ws.cell(row=grand_row + 3, column=1, value="※ 追加の立替が判明した場合はこの表を更新すること。")

    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# 収支報告書（保護者向け・昨年様式踏襲）
# ---------------------------------------------------------------------------


def write_camp_report_xlsx(path: str | Path = None, camp: model.Camp = None) -> Path:
    if camp is None:
        camp = model.load()
    if path is None:
        path = OUTPUT_DIR / f"{camp.id}_夏合宿収支報告.xlsx"
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    inc = model.income(camp)
    bill = model.hotel_bill(camp)
    bal = model.balance(camp)

    full_select = [l for l in inc["lines"] if l.is_full_time and l.person.role == "選手"]
    full_mgr = [l for l in inc["lines"] if l.is_full_time and l.person.role == "マネージャー"]
    late = [l for l in inc["lines"] if not l.is_full_time]
    total_selects = sum(1 for l in inc["lines"] if l.person.role == "選手")
    total_people = len(inc["lines"])

    b_expenses = [e for e in camp.expenses if e.category == "B"]
    c_expenses = [e for e in camp.expenses if e.category == "C"]

    wb = Workbook()
    ws = wb.active
    ws.title = "収支報告"
    _autosize(ws, [30, 14, 12, 16, 14])

    r = 1
    ws.cell(row=r, column=1, value="ラグビー部保護者各位").font = BOLD
    r += 1
    ws.cell(row=r, column=1, value="日付:")
    ws.cell(row=r, column=2, value=date.today()).number_format = "yyyy年m月d日"
    r += 1
    ws.cell(row=r, column=1, value="差出人: 駒澤大学高等学校 校長 井上誠二 / ラグビー部顧問 畠山和真・占部涼也")
    r += 2
    ws.cell(row=r, column=1, value=f"{camp.name} 収支報告").font = TITLE_FONT
    r += 1
    ws.cell(
        row=r,
        column=1,
        value=(
            "夏合宿（"
            f"{date_ja(camp.start)}〜{date_ja(camp.end)}・{camp.venue}"
            "）にあたりましては、格別のご協力を賜り誠にありがとうございました。"
            "下記の通り収支がまとまりましたのでご報告いたします。"
        ),
    )
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    ws.row_dimensions[r].height = 30
    ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    r += 2

    # ---- 収入 -------------------------------------------------------------
    income_header_row = r
    ws.cell(row=r, column=1, value="【収入】").font = BOLD
    r += 1
    for c, h in enumerate(["費目", "単価", "人数", "金額"], start=1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = BOLD
        cell.fill = HEADER_FILL
        cell.border = BOX
    r += 1
    income_first_row = r

    def income_row(label, unit, count, pending=False):
        nonlocal r
        ws.cell(row=r, column=1, value=label)
        u = ws.cell(row=r, column=2, value=unit)
        u.number_format = YEN_FMT
        cnt = ws.cell(row=r, column=3, value=count)
        amt = ws.cell(row=r, column=4, value=f"=B{r}*C{r}")
        amt.number_format = YEN_FMT
        for c in range(1, 5):
            ws.cell(row=r, column=c).border = BOX
        if pending:
            u.fill = YELLOW
            cnt.fill = YELLOW
        r += 1

    income_row("選手参加費（全日参加）", 79090, len(full_select))
    income_row("マネージャー参加費（全日参加）", 78060, len(full_mgr))
    for l in late:
        income_row(f"途中参加者（{l.person.name}・{l.person.role}）", l.total, 1)
    # 顧問（引率）の宿泊・交通・昼食代: 生徒徴収とは別財源から充当（未解決の論点1）。金額未確定。
    income_row("顧問宿泊費・交通費・昼食代（別財源）", 0, 1, pending=True)

    income_total_row = r
    ws.cell(row=r, column=1, value="収入合計").font = BOLD
    income_total_cell = ws.cell(row=r, column=4, value=f"=SUM(D{income_first_row}:D{r - 1})")
    income_total_cell.font = BOLD
    income_total_cell.number_format = YEN_FMT
    income_total_cell.fill = HEADER_FILL
    for c in range(1, 5):
        ws.cell(row=r, column=c).border = BOX
    r += 2

    # ---- 支出 -------------------------------------------------------------
    ws.cell(row=r, column=1, value="【支出】").font = BOLD
    r += 1
    for c, h in enumerate(["費目", "単価", "数量", "金額"], start=1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = BOLD
        cell.fill = HEADER_FILL
        cell.border = BOX
    r += 1
    expense_first_row = r

    def expense_row(label, unit, qty, pending=False, note=None):
        nonlocal r
        ws.cell(row=r, column=1, value=label)
        u = ws.cell(row=r, column=2, value=unit)
        u.number_format = YEN_FMT
        q = ws.cell(row=r, column=3, value=qty)
        amt = ws.cell(row=r, column=4, value=f"=B{r}*C{r}")
        amt.number_format = YEN_FMT
        for c in range(1, 5):
            ws.cell(row=r, column=c).border = BOX
        if pending:
            u.fill = YELLOW
        if note:
            ws.cell(row=r, column=5, value=note)
        row_ref = r
        r += 1
        return row_ref

    expense_row("ホテル宿泊代（1泊3食）", bill["items"][0]["unit"], bill["items"][0]["qty"])
    expense_row("増（追加）昼食代", bill["items"][1]["unit"], bill["items"][1]["qty"])
    expense_row("BBQ代", bill["items"][2]["unit"], bill["items"][2]["qty"])
    expense_row("グラウンド使用料", bill["items"][3]["unit"], bill["items"][3]["qty"])
    expense_row("グラウンド使用料", bill["items"][4]["unit"], bill["items"][4]["qty"])
    gap = camp.hotel_invoice_total - bill["total"]
    expense_row(
        "ホテル請求書との差異（増朝食・増夕食・増宿泊等）※1",
        gap,
        1,
        note="簡易再構成モデルでは内訳化していない単発追加分",
    )
    misc_row = expense_row("雑費（合宿中の経費）※2", sum(e.amount for e in c_expenses), 1)
    bus_row = expense_row("バス代（国際興業・見積額）※3", camp.bus["quote"], 1)
    coach_row = expense_row(
        "コーチ宿泊代・交通費・謝礼", 0, 1, pending=True, note="金額未確定（畠山先生に確認中）"
    )
    reserve_row = expense_row("予備費未使用返金分※4", camp.collection["reserve"], total_people)

    expense_total_row = r
    ws.cell(row=r, column=1, value="支出合計").font = BOLD
    expense_total_cell = ws.cell(row=r, column=4, value=f"=SUM(D{expense_first_row}:D{r - 1})")
    expense_total_cell.font = BOLD
    expense_total_cell.number_format = YEN_FMT
    expense_total_cell.fill = HEADER_FILL
    for c in range(1, 5):
        ws.cell(row=r, column=c).border = BOX
    r += 2

    # ---- 収支差額・バス代差額・合宿残金 ------------------------------------
    diff_row = r
    ws.cell(row=r, column=1, value="収支差額（収入－支出）").font = BOLD
    diff_cell = ws.cell(row=r, column=4, value=f"=D{income_total_row}-D{expense_total_row}")
    diff_cell.font = BOLD
    diff_cell.number_format = YEN_FMT
    r += 1

    bus_diff_row = r
    ws.cell(row=r, column=1, value="バス代差額（有料道路代・駐車場代など後日請求分）")
    bus_pending = ws.cell(row=r, column=4, value=0)
    bus_pending.number_format = YEN_FMT
    bus_pending.fill = YELLOW
    ws.cell(row=r, column=5, value="未確定（実費請求が届き次第、マイナスとして入力）")
    r += 1

    remain_row = r
    ws.cell(row=r, column=1, value="合宿残金").font = BOLD
    remain_cell = ws.cell(row=r, column=4, value=f"=D{diff_row}-D{bus_diff_row}")
    remain_cell.font = BOLD
    remain_cell.number_format = YEN_FMT
    remain_cell.fill = HEADER_FILL
    r += 2

    # ---- 注釈 ---------------------------------------------------------------
    notes = [
        "※1 ホテル請求書合計は菅平ホテル発行の請求書（支払済）の実額。上の内訳4行の合計との差額をこの行で吸収しています。",
        "※2 雑費は「合宿中使用経費」シートの領収書付き支出（会計分類C）の合計です。会計分類B（父母会予算で精算するもの）は含みません。",
        "※3 バス代は国際興業への見積額。有料道路代・駐車場代は合宿後に実費請求されるため、届き次第「バス代差額」に反映してください。",
        "※4 予備費は不参加者を除く全参加者から@" f"{camp.collection['reserve']:,}円徴収し、使用しなかったため全額返金します。",
        "※ コーチ宿泊代・交通費・謝礼（黄色セル）は金額未確定です。判明次第このセルに入力すれば、支出合計以下は自動で再計算されます。",
        "※ 合宿残金は返金せず、合宿後のプロテイン代等、部の活動費に充当します。",
        "※ 対戦校への手土産・メディカル費等（会計分類B）は父母会予算から別途精算するため、上記の合宿費会計には含めていません。詳細は「父母会予算精算」シートを参照してください。",
    ]
    for note in notes:
        ws.cell(row=r, column=1, value=note)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True)
        r += 1

    # ---- シート2: 合宿中使用経費（雑費一次記録） ----------------------------
    ws2 = wb.create_sheet("合宿中使用経費")
    _autosize(ws2, [12, 10, 24, 30, 12, 12, 10, 24])
    ws2.cell(row=1, column=1, value="合宿中使用経費（雑費一次記録・会計分類C）").font = TITLE_FONT
    headers2 = ["日付", "雑費番号", "支払先", "内容", "金額", "支払者", "精算", "備考"]
    for c, h in enumerate(headers2, start=1):
        cell = ws2.cell(row=3, column=c, value=h)
        cell.font = BOLD
        cell.fill = HEADER_FILL
        cell.border = BOX
    rr = 4
    for i, e in enumerate(c_expenses, start=1):
        ws2.cell(row=rr, column=1, value=e.date).number_format = "m/d"
        ws2.cell(row=rr, column=2, value=i)
        ws2.cell(row=rr, column=3, value=e.vendor)
        ws2.cell(row=rr, column=4, value=e.description)
        amt = ws2.cell(row=rr, column=5, value=e.amount)
        amt.number_format = YEN_FMT
        ws2.cell(row=rr, column=6, value=e.payer)
        ws2.cell(row=rr, column=7, value="済" if e.is_settled else "未")
        ws2.cell(row=rr, column=8, value=e.note)
        for c in range(1, 9):
            ws2.cell(row=rr, column=c).border = BOX
        rr += 1
    ws2.cell(row=rr, column=4, value="雑費 合計").font = BOLD
    total2 = ws2.cell(row=rr, column=5, value=f"=SUM(E4:E{rr - 1})")
    total2.font = BOLD
    total2.number_format = YEN_FMT
    ws2.cell(row=rr + 2, column=1, value=f"※ 今年度は{len(c_expenses)}件。昨年度は雑費36件の記録があり、比較して記録漏れがないか確認してください。")

    # ---- シート3: 父母会予算で精算するもの（会計分類B） --------------------
    ws3 = wb.create_sheet("父母会予算精算")
    _autosize(ws3, [12, 24, 30, 12, 12, 12, 24])
    ws3.cell(row=1, column=1, value="合宿中の支出で父母会予算（会計分類B）で精算するもの").font = TITLE_FONT
    headers3 = ["日付", "支払先", "内容", "金額", "支払者", "精算", "備考"]
    for c, h in enumerate(headers3, start=1):
        cell = ws3.cell(row=3, column=c, value=h)
        cell.font = BOLD
        cell.fill = HEADER_FILL
        cell.border = BOX
    rr = 4
    for e in b_expenses:
        ws3.cell(row=rr, column=1, value=e.date).number_format = "m/d"
        ws3.cell(row=rr, column=2, value=e.vendor)
        ws3.cell(row=rr, column=3, value=e.description)
        amt = ws3.cell(row=rr, column=4, value=e.amount)
        amt.number_format = YEN_FMT
        ws3.cell(row=rr, column=5, value=e.payer)
        ws3.cell(row=rr, column=6, value="済" if e.is_settled else "未")
        ws3.cell(row=rr, column=7, value=e.note)
        for c in range(1, 8):
            ws3.cell(row=rr, column=c).border = BOX
        rr += 1
    ws3.cell(row=rr, column=3, value="父母会予算 合計").font = BOLD
    total3 = ws3.cell(row=rr, column=4, value=f"=SUM(D4:D{rr - 1})")
    total3.font = BOLD
    total3.number_format = YEN_FMT
    ws3.cell(
        row=rr + 2,
        column=1,
        value="※ 対戦校への手土産代・メディカル費等。合宿費（会計分類C）の徴収額には混ぜず、父母会予算から別途精算します。",
    )

    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# 人数表（ホテル連絡用）
# ---------------------------------------------------------------------------


def write_headcount_xlsx(path: str | Path = None, camp: model.Camp = None) -> Path:
    if camp is None:
        camp = model.load()
    if path is None:
        path = OUTPUT_DIR / f"{camp.id}_日別人数表.xlsx"
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    hc = model.headcount(camp)

    wb = Workbook()
    ws = wb.active
    ws.title = "日別人数表"
    _autosize(ws, [16, 10, 10, 10, 10, 14])

    ws["A1"] = f"{camp.name}（{camp.venue}）日別人数表"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "作成日:"
    ws["B2"] = date.today()
    ws["B2"].number_format = "yyyy/mm/dd"

    header_row = 4
    headers = ["日付", "朝食", "昼食", "夕食", "宿泊", "申告済み人数"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=c, value=h)
        cell.font = BOLD
        cell.fill = HEADER_FILL
        cell.border = BOX

    row = header_row + 1
    first_row = row
    for d in camp.date_range():
        v = hc["per_date"][d]
        ws.cell(row=row, column=1, value=date_ja(d))
        for c, key in zip((2, 3, 4, 5), ("breakfast", "lunch", "dinner", "lodging")):
            val = v[key]
            cell = ws.cell(row=row, column=c, value=(val if val is not None else "－"))
            if val is not None:
                cell.number_format = YEN_FMT
            cell.alignment = Alignment(horizontal="center")
        submitted = camp.submitted_headcount.get(d)
        scell = ws.cell(row=row, column=6, value=submitted if submitted is not None else "－")
        scell.alignment = Alignment(horizontal="center")
        if submitted is not None and v["present"] != submitted:
            scell.fill = YELLOW
        for c in range(1, 7):
            ws.cell(row=row, column=c).border = BOX
        row += 1

    total_row = row
    ws.cell(row=total_row, column=1, value="合計（延べ人数）").font = BOLD
    for c, key in zip((2, 3, 4, 5), ("breakfast", "lunch", "dinner", "lodging")):
        col = get_column_letter(c)
        cell = ws.cell(row=total_row, column=c, value=f"=SUM({col}{first_row}:{col}{row - 1})")
        cell.font = BOLD
        cell.number_format = YEN_FMT
    for c in range(1, 7):
        ws.cell(row=total_row, column=c).border = BOX

    ws.cell(
        row=total_row + 2,
        column=1,
        value="※ 「宿泊」列＝その夜の夕食人数。合宿初日は朝食なし、最終日は夕食・宿泊なしとして計算しています。",
    )
    ws.cell(
        row=total_row + 3,
        column=1,
        value="※ 黄色セルは畠山先生の申告済み人数と名簿からの計算人数が一致していない日です。check コマンドの結果もあわせて確認してください。",
    )

    wb.save(path)
    return path
